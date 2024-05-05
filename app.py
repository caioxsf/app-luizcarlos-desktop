import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook


# aparencia do sistema 
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema Planilha Luiz Carlos")
        self.geometry("700x500")

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff']).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "Sistema"], command=self.change_apm).place(x=50, y=460)

    def todo_sistema (self):
        frame = ctk.CTkFrame(self, width=750, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema Planilha Luiz Carlos", font=("Century Gothic bold", 24), text_color="#fff").place(x=190, y=10)
        span = ctk.CTkLabel(self, text="Preencha todos os campos se necessário!", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)
        span = ctk.CTkLabel(self, text="(Frete, Peso, Adiantamento e Salário precisam ser preenchidos com 0)", font=("Century Gothic bold", 13), text_color=["#000", "#808080"]).place(x=50, y=90)
        plan = pathlib.Path("Planilha.xlsx")   

        #inicializando as var de 0
        global valor_total_viagens, comissao_total, adiantamento_total
        valor_total_viagens = 0
        comissao_total = 0
        adiantamento_total = 0

        if plan.exists():
            #se caso ainda nao exista a planilha, ela cria uma
            #se ja existir, ela pega os valores totais e atualiza nas variaveis que zeraram
            plan = openpyxl.load_workbook('Planilha.xlsx')
            folha = plan.active

            for row in folha.iter_rows(min_row=2, max_col=16, max_row=folha.max_row, values_only=True):
                valor_total_viagens += row[13] or 0  #soma o valor da coluna 14 (índice 13)
                comissao_total += row[14] or 0       #soma o valor da coluna 15 (índice 14)
                adiantamento_total += row[15] or 0   #soma o valor da coluna 16 (índice 15)

            plan.close()
        else:
            plan=Workbook()
            folha=plan.active
            folha['A1']="Data"
            folha['B1']="Origem"
            folha['C1']="Destino"
            folha['D1']="Material"
            folha['E1']="Motorista"
            folha['F1']="Transportadora"
            folha['G1']="Frete"
            folha['H1']="Peso"
            folha['I1']="Valor total"
            folha['J1']="Comissao"
            folha['K1']="Data adiantamento"
            folha['L1']="Adiantamento"

            folha['N1']="Valor total viagens"
            folha['O1']="Valor total comissões"
            folha['P1']="Valor total do adiantamento"

            folha['Q1']="Salario motorista"

            plan.save("Planilha.xlsx")

        def submit():
            global data_value, origem_value, destino_value, frete_value, material_value, peso_value, motorista_value, transportadora_value, adiantamento_value, data_adiatamento_value, valor_total_viagens, comissao_total, adiantamento_total

            #pegando os dados dos entrys(inputs)
            data = data_entry.get()
            origem = origem_entry.get()
            destino = destino_entry.get()
            frete = frete_entry.get()
            material = material_entry.get()
            peso = peso_entry.get()
            motorista = motorista_entry.get()
            transportadora = transportadora_entry.get()
            adiantamento = adiantamento_entry.get()
            data_adiatamento = data_adiatamento_entry.get()

            #passando de string pra float 
            frete = float(frete_entry.get())
            peso = float(peso_entry.get())
            adiantamento = float(adiantamento_entry.get())

            #calculandoa valor total e comissao
            valor_total = frete * peso
            comissao = valor_total * 8 / 100

            #calculando o valor total da comissao e das viagens e adiantamento
            valor_total_viagens = valor_total + valor_total_viagens
            comissao_total = comissao + comissao_total
            adiantamento_total = adiantamento + adiantamento_total

            #arredondando para duas casas decimais
            valor_total = round(valor_total, 2)
            comissao = round(comissao, 2)

            #salario motorista
            salario = comissao_total + 2500 - adiantamento_total

            plan = openpyxl.load_workbook('Planilha.xlsx')
            folha = plan.active
            folha.cell(column=1, row=folha.max_row+1, value=data)
            folha.cell(column=2, row=folha.max_row, value=origem)
            folha.cell(column=3, row=folha.max_row, value=destino)
            folha.cell(column=4, row=folha.max_row, value=material)
            folha.cell(column=5, row=folha.max_row, value=motorista)
            folha.cell(column=6, row=folha.max_row, value=transportadora)
            folha.cell(column=7, row=folha.max_row, value=frete)
            folha.cell(column=8, row=folha.max_row, value=peso)
            folha.cell(column=9, row=folha.max_row, value=valor_total)
            folha.cell(column=10, row=folha.max_row, value=comissao)
            folha.cell(column=11, row=folha.max_row, value=data_adiatamento)
            folha.cell(column=12, row=folha.max_row, value=adiantamento)

            #colunas e linhas dos valores totais e a receber
            folha.cell(column=14, row=2, value=valor_total_viagens)
            folha.cell(column=15, row=2, value=comissao_total)
            folha.cell(column=16, row=2, value=adiantamento_total)

            #salario motorista
            folha.cell(column=17, row=2, value=salario)
            
            plan.save(r"Planilha.xlsx")
            messagebox.showinfo("Sistema", "Dados salvos com sucesso!")

        def clear():
            data_value.set("")
            origem_value.set("")
            destino_value.set("")
            frete_value.set("")
            material_value.set("")
            peso_value.set("")
            motorista_value.set("")
            transportadora_value.set("")
            adiantamento_value.set("")
            data_adiatamento_value("")

        #texts variables
        data_value = StringVar()
        origem_value = StringVar()
        destino_value = StringVar()
        frete_value = StringVar()
        material_value = StringVar()
        peso_value = StringVar()
        motorista_value = StringVar()
        transportadora_value = StringVar()
        adiantamento_value = StringVar()
        data_adiatamento_value = StringVar()

        #entrys (input)
        data_entry = ctk.CTkEntry(self, width=135, textvariable=data_value, font=("Century Gothic bold", 16), fg_color="transparent")
        origem_entry = ctk.CTkEntry(self, width=135, textvariable=origem_value, font=("Century Gothic bold", 16), fg_color="transparent")
        destino_entry = ctk.CTkEntry(self, width=135, textvariable=destino_value, font=("Century Gothic bold", 16), fg_color="transparent")
        frete_entry = ctk.CTkEntry(self, width=135, textvariable=frete_value, font=("Century Gothic bold", 16), fg_color="transparent")
        material_entry = ctk.CTkEntry(self, width=135, textvariable=material_value, font=("Century Gothic bold", 16), fg_color="transparent")
        peso_entry = ctk.CTkEntry(self, width=135, textvariable=peso_value, font=("Century Gothic bold", 16), fg_color="transparent")
        motorista_entry = ctk.CTkEntry(self, width=135, textvariable=motorista_value, font=("Century Gothic bold", 16), fg_color="transparent")
        transportadora_entry = ctk.CTkEntry(self, width=135, textvariable=transportadora_value, font=("Century Gothic bold", 16), fg_color="transparent")
        adiantamento_entry = ctk.CTkEntry(self, width=135, textvariable=adiantamento_value, font=("Century Gothic bold", 16), fg_color="transparent")
        data_adiatamento_entry = ctk.CTkEntry(self, width=135, textvariable=data_adiatamento_value, font=("Century Gothic bold", 16), fg_color="transparent")

        #labels
        lb_data = ctk.CTkLabel(self, text="Data", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_origem = ctk.CTkLabel(self, text="Origem", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_destino = ctk.CTkLabel(self, text="Destino", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_frete = ctk.CTkLabel(self, text="Frete", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_material = ctk.CTkLabel(self, text="Material", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_peso = ctk.CTkLabel(self, text="Peso", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_motorista = ctk.CTkLabel(self, text="Motorista", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_transportadora = ctk.CTkLabel(self, text="Transportadora", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_adiatamento = ctk.CTkLabel(self, text="Adiatamento", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_data_adiantamento = ctk.CTkLabel(self, text="Data adintamento", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=230, y=460)
        btn_clear = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=380, y=460)

        #posicionando os elementos na tela
        lb_data.place(x=50, y=120)
        data_entry.place(x=50, y=150)

        lb_origem.place(x=200, y=120)
        origem_entry.place(x=200, y=150)

        lb_destino.place(x=350, y=120)
        destino_entry.place(x=350, y=150)

        lb_frete.place(x=500, y=120)
        frete_entry.place(x=500, y=150)

        lb_material.place(x=50, y=190)
        material_entry.place(x=50, y=220)

        lb_peso.place(x=200, y=190)
        peso_entry.place(x=200, y=220)

        lb_motorista.place(x=350, y=190)
        motorista_entry.place(x=350, y=220)

        lb_transportadora.place(x=500, y=190)
        transportadora_entry.place(x=500, y=220)

        lb_adiatamento.place(x=50, y=260)
        adiantamento_entry.place(x=50, y=290)

        lb_data_adiantamento.place(x=200, y=260)
        data_adiatamento_entry.place(x=200, y=290)

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__ == "__main__":
    app = App()
    app.mainloop()